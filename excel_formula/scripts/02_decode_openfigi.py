"""
Stage 0b: read all unique BBG/FIGI IDs from
data/Excel5b_UniqueTickers_ForDecode.xlsx (output of
01_process_excel5.py), decode them via the OpenFIGI public API, and write
one formatted Excel per underlying + a master file into
data/decoded/.

Usage
-----
    Set OPENFIGI_API_KEY in the project's .env (loaded automatically), or
    export it in the shell, then:
        python excel_formula/scripts/02_decode_openfigi.py

Resume: just re-run the same command. Already-decoded FIGIs are loaded
from decoded/_checkpoint.json and skipped automatically -- if the
checkpoint already covers every ID (as it does after the initial run),
no API key is needed and no requests are made.

Rate limits (with free API key):
    25 requests per 6-second window, 100 FIGIs per request
    146k IDs ÷ 100 = ~1,462 requests ≈ ~6 minutes

Output feeds excel_formula/scripts/03_filter_contracts.py.
"""

import os, sys, re, json, time, math, logging
from datetime import timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import requests
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"

load_dotenv(dotenv_path=ROOT / ".env")

INPUT_FILE    = DATA_DIR / "Excel5b_UniqueTickers_ForDecode.xlsx"
OUTPUT_DIR    = DATA_DIR / "decoded"
CHECKPOINT    = os.path.join(OUTPUT_DIR, "_checkpoint.json")
LOG_FILE      = os.path.join(OUTPUT_DIR, "decode_run.log")
MASTER_FILE   = os.path.join(OUTPUT_DIR, "ALL_decoded.xlsx")
SKIP_SHEETS   = {"INDEX", "HOW_TO_USE"}

OPENFIGI_URL  = "https://api.openfigi.com/v3/mapping"
BATCH_SIZE    = 100      # max with API key
SAVE_EVERY    = 50       # checkpoint every N batches (~5,000 FIGIs)
# Conservative: 25 req per 6 s = 4.17 req/s → 0.25 s is safe default
# We also read ratelimit-remaining from headers and back off dynamically
DEFAULT_SLEEP = 0.26

# ── COLORS ────────────────────────────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2E75B6"
WHITE  = "FFFFFF"
LGREY  = "F2F2F2"
GREEN  = "E2EFDA"
YELLOW = "FFF2CC"
RED    = "FFDCD8"

def fill(c):     return PatternFill("solid", fgColor=c)
def hfont(sz=9): return Font(name="Arial", bold=True,  size=sz, color="FFFFFF")
def dfont(sz=9): return Font(name="Arial", bold=False, size=sz, color="000000")
def mfont(sz=8): return Font(name="Courier New", bold=False, size=sz, color="000080")
def center():    return Alignment(horizontal="center", vertical="center")
def left():      return Alignment(horizontal="left",   vertical="center")
def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ── LOGGING ───────────────────────────────────────────────────────────────────
def setup_logging() -> logging.Logger:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    fmt = "%(asctime)s  %(levelname)-7s  %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"
    logger = logging.getLogger("openfigi")
    logger.setLevel(logging.DEBUG)
    # Console handler — INFO and above
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter(fmt, datefmt))
    # File handler — DEBUG and above (keeps full detail)
    fh = logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(fmt, datefmt))
    logger.addHandler(ch)
    logger.addHandler(fh)
    return logger

log = setup_logging()


# ── OPTION DESCRIPTION PARSER ─────────────────────────────────────────────────
# securityDescription format: "MSTR US 01/17/25 C140"
_OPT_RE = re.compile(
    r"^(?P<under>[A-Z0-9]+)\s+"
    r"(?P<exch>[A-Z]+)\s+"
    r"(?P<mm>\d{2})/(?P<dd>\d{2})/(?P<yy>\d{2})\s+"
    r"(?P<pc>[CP])(?P<strike>[\d.]+)$",
    re.IGNORECASE
)

def parse_description(desc: str) -> dict:
    if not desc:
        return {}
    m = _OPT_RE.match(desc.strip())
    if not m:
        return {}
    yr = int(m.group("yy"))
    expiry = f"{2000 + yr}-{m.group('mm')}-{m.group('dd')}"
    return {
        "underlying": m.group("under").upper(),
        "exchange":   m.group("exch").upper(),
        "expiry":     expiry,
        "opt_type":   "Call" if m.group("pc").upper() == "C" else "Put",
        "strike":     float(m.group("strike")),
    }


# ── OPENFIGI API ──────────────────────────────────────────────────────────────
def get_api_key() -> str:
    key = os.environ.get("OPENFIGI_API_KEY", "").strip()
    if not key:
        log.error("OPENFIGI_API_KEY not set.")
        log.error("Get a free key at: https://www.openfigi.com/api#apikey")
        log.error("Then run:  export OPENFIGI_API_KEY='your-key-here'")
        sys.exit(1)
    return key


def query_openfigi(batch: List[str], api_key: str,
                   retries: int = 5) -> Tuple[List[Optional[dict]], float]:
    """
    POST a batch of ≤100 FIGIs to OpenFIGI.
    Returns (results, suggested_sleep_seconds).
    results[i] is None for unresolved FIGIs.
    """
    payload = [{"idType": "ID_BB_GLOBAL", "idValue": figi} for figi in batch]
    headers = {
        "Content-Type":    "application/json",
        "X-OPENFIGI-APIKEY": api_key,
    }
    sleep_hint = DEFAULT_SLEEP

    for attempt in range(1, retries + 1):
        try:
            r = requests.post(OPENFIGI_URL, json=payload, headers=headers, timeout=30)

            # Parse rate-limit headers for adaptive throttling
            remaining = r.headers.get("ratelimit-remaining")
            reset_secs = r.headers.get("ratelimit-reset")
            if remaining is not None and reset_secs is not None:
                remaining = int(remaining)
                reset_secs = float(reset_secs)
                log.debug(f"  rate-limit: remaining={remaining}, reset_in={reset_secs:.1f}s")
                if remaining <= 2 and reset_secs > 0:
                    # Almost out of quota — wait for window to reset
                    sleep_hint = reset_secs + 0.2
                elif remaining <= 5:
                    sleep_hint = max(DEFAULT_SLEEP, reset_secs / max(remaining, 1))
                else:
                    sleep_hint = DEFAULT_SLEEP

            if r.status_code == 429:
                wait = float(reset_secs) + 1 if reset_secs else 10 * attempt
                log.warning(f"  429 rate-limited — sleeping {wait:.1f}s (attempt {attempt})")
                time.sleep(wait)
                continue

            if r.status_code != 200:
                log.warning(f"  HTTP {r.status_code} attempt {attempt}: {r.text[:300]}")
                time.sleep(5 * attempt)
                continue

            out = []
            for item in r.json():
                if "error" in item:
                    out.append(None)
                else:
                    data = item.get("data", [])
                    out.append(data[0] if data else None)
            return out, sleep_hint

        except requests.exceptions.Timeout:
            log.warning(f"  Timeout on attempt {attempt} — retrying in {5*attempt}s")
            time.sleep(5 * attempt)
        except Exception as e:
            log.warning(f"  Request error attempt {attempt}: {e}")
            time.sleep(5 * attempt)

    log.error(f"  All {retries} attempts failed for batch starting {batch[0]}")
    return [None] * len(batch), DEFAULT_SLEEP


# ── CHECKPOINT ────────────────────────────────────────────────────────────────
def load_checkpoint() -> dict:
    if os.path.exists(CHECKPOINT):
        try:
            with open(CHECKPOINT) as f:
                data = json.load(f)
            log.info(f"Checkpoint loaded: {len(data):,} already-decoded FIGIs")
            return data
        except Exception as e:
            log.warning(f"Could not load checkpoint ({e}) — starting fresh")
    return {}

def save_checkpoint(cache: dict):
    tmp = CHECKPOINT + ".tmp"
    with open(tmp, "w") as f:
        json.dump(cache, f)
    os.replace(tmp, CHECKPOINT)   # atomic replace
    log.debug(f"Checkpoint saved: {len(cache):,} entries")


# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
COLUMNS    = ["figi", "raw_id", "description", "ticker",
              "underlying", "exchange", "expiry", "strike",
              "opt_type", "security_type", "market_sector"]
COL_WIDTHS = [18, 26, 30, 26, 14, 10, 14, 10, 8, 20, 16]
HEADERS    = ["FIGI", "Raw BBG ID", "Security Description", "Ticker",
              "Underlying", "Exchange", "Expiry Date", "Strike",
              "Opt Type", "Security Type", "Market Sector"]

def _safe_stats(df: pd.DataFrame) -> str:
    parts = []
    if "opt_type" in df.columns:
        calls = (df["opt_type"] == "Call").sum()
        puts  = (df["opt_type"] == "Put").sum()
        parts.append(f"{calls:,} calls  {puts:,} puts")
    if "expiry" in df.columns:
        parts.append(f"Unique expiries: {df['expiry'].nunique():,}")
    if "strike" in df.columns:
        s = df["strike"].dropna()
        if len(s):
            parts.append(f"Strike range: {s.min():.0f}–{s.max():.0f}")
    return "  |  ".join(parts)

def df_to_sheet(ws, df: pd.DataFrame, sheet_label: str):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    last_col = get_column_letter(len(COLUMNS))

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"{sheet_label} — {len(df):,} decoded contracts"
    c.font = hfont(11); c.fill = fill(NAVY); c.alignment = left()
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = "Decoded via OpenFIGI API  |  " + _safe_stats(df)
    c.font = Font(name="Arial", size=8, italic=True, color="FFFFFF")
    c.fill = fill(BLUE); c.alignment = left()
    ws.row_dimensions[2].height = 14

    for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.font = hfont(9); c.fill = fill(BLUE); c.alignment = center()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 16

    for ri, (_, row) in enumerate(df.iterrows()):
        r   = ri + 4
        bg  = WHITE if ri % 2 == 0 else LGREY
        for ci, col in enumerate(COLUMNS, 1):
            val = row.get(col, "")
            val = "" if pd.isna(val) else val
            c   = ws.cell(row=r, column=ci, value=val)
            c.font = mfont(8) if ci == 1 else dfont(9)
            if ci == 7:           c.fill = fill(GREEN)
            elif ci == 8:         c.fill = fill(YELLOW)
            elif val == "":       c.fill = fill(RED)
            else:                 c.fill = fill(bg)
            c.alignment = center() if ci in [1, 6, 7, 8, 9, 10, 11] else left()
            c.border = thin()
        ws.row_dimensions[r].height = 13


def write_decoded_excel(path: str, df: pd.DataFrame, sheet_name: str):
    wb = Workbook()
    wb.remove(wb.active)
    df_to_sheet(wb.create_sheet(sheet_name[:31]), df, sheet_name)
    wb.save(path)


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    run_start = time.time()
    log.info("=" * 60)
    log.info("OpenFIGI decode — starting")
    log.info(f"Log file: {LOG_FILE}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ── Read FIGIs from Excel5b ───────────────────────────────────────────────
    log.info(f"Reading {INPUT_FILE} …")
    wb = load_workbook(INPUT_FILE, data_only=True)
    sheet_figis: Dict[str, List[str]] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        ids = [
            str(row[0]).strip()
            for row in ws.iter_rows(min_row=4, values_only=True)
            if row[0] and str(row[0]).strip().startswith("BBG")
        ]
        if ids:
            sheet_figis[sheet_name] = ids
            log.info(f"  {sheet_name:8s}: {len(ids):>6,} FIGIs")

    total_figis = sum(len(v) for v in sheet_figis.values())
    log.info(f"Total FIGIs: {total_figis:,}")

    def bare(raw: str) -> str:
        return raw.split()[0]

    all_items: List[Tuple[str, str, str]] = [
        (underlying, raw, bare(raw))
        for underlying, ids in sheet_figis.items()
        for raw in ids
    ]

    # ── Load checkpoint ───────────────────────────────────────────────────────
    cache: Dict[str, dict] = load_checkpoint()
    pending = [item for item in all_items if item[2] not in cache]
    log.info(f"Already decoded: {total_figis - len(pending):,}  |  Pending: {len(pending):,}")

    # ── API decode loop ───────────────────────────────────────────────────────
    if pending:
        api_key = get_api_key()
        total_batches = math.ceil(len(pending) / BATCH_SIZE)
        log.info(f"Sending {total_batches:,} requests  "
                 f"({BATCH_SIZE} FIGIs each)  …")

        api_failures   = 0
        decoded_this_run = 0
        sleep_secs     = DEFAULT_SLEEP

        for b_idx in range(total_batches):
            chunk    = pending[b_idx * BATCH_SIZE : (b_idx + 1) * BATCH_SIZE]
            bare_ids = [item[2] for item in chunk]

            results, sleep_secs = query_openfigi(bare_ids, api_key)

            for (underlying, raw, figi), res in zip(chunk, results):
                if res is None:
                    api_failures += 1
                    cache[figi] = {"figi": figi, "raw_id": raw}
                    log.debug(f"  NO_DATA  {figi}")
                else:
                    desc   = res.get("securityDescription") or res.get("name", "")
                    parsed = parse_description(desc)
                    cache[figi] = {
                        "figi":          figi,
                        "raw_id":        raw,
                        "description":   desc,
                        "ticker":        res.get("ticker", ""),
                        "security_type": res.get("securityType", ""),
                        "market_sector": res.get("marketSector", ""),
                        **parsed,
                    }
                    decoded_this_run += 1
                    log.debug(f"  OK  {figi}  →  {desc}")

            # ── Progress line every 25 batches ───────────────────────────────
            if (b_idx + 1) % 25 == 0 or b_idx == total_batches - 1:
                pct      = (b_idx + 1) / total_batches * 100
                done_n   = min((b_idx + 1) * BATCH_SIZE, len(pending))
                elapsed  = time.time() - run_start
                rate     = done_n / elapsed if elapsed > 0 else 1
                eta_secs = (len(pending) - done_n) / rate if rate else 0
                eta_str  = str(timedelta(seconds=int(eta_secs)))
                log.info(
                    f"  [{pct:5.1f}%]  batch {b_idx+1:,}/{total_batches:,}  "
                    f"decoded {done_n:,}/{len(pending):,}  "
                    f"failures {api_failures:,}  ETA {eta_str}"
                )

            # ── Periodic checkpoint ───────────────────────────────────────────
            if (b_idx + 1) % SAVE_EVERY == 0:
                save_checkpoint(cache)
                log.info(f"  Checkpoint saved  ({len(cache):,} total entries)")

            time.sleep(sleep_secs)

        save_checkpoint(cache)
        elapsed_total = time.time() - run_start
        log.info(f"API loop complete in {timedelta(seconds=int(elapsed_total))}  "
                 f"— decoded {decoded_this_run:,}  failures {api_failures:,}")

    # ── Build DataFrame ───────────────────────────────────────────────────────
    log.info("Building DataFrames …")
    rows = []
    for underlying, raw, figi in all_items:
        entry = dict(cache.get(figi, {"figi": figi, "raw_id": raw}))
        entry["_underlying"] = underlying
        rows.append(entry)

    df_all = pd.DataFrame(rows)
    for col in COLUMNS:
        if col not in df_all.columns:
            df_all[col] = ""

    df_all = df_all.sort_values(["_underlying", "expiry", "strike"], na_position="last")

    # ── Per-underlying Excel files ────────────────────────────────────────────
    log.info(f"Writing decoded files to  {OUTPUT_DIR}/  …")
    for underlying in sheet_figis:
        sub = df_all[df_all["_underlying"] == underlying][COLUMNS].copy()
        if sub.empty:
            continue
        path = os.path.join(OUTPUT_DIR, f"{underlying}_decoded.xlsx")
        write_decoded_excel(path, sub, f"{underlying} Options — Decoded")
        log.info(f"  ✓ {path}  ({len(sub):,} rows)")

    # ── Master Excel (summary + per-underlying sheets) ────────────────────────
    log.info(f"Writing master file: {MASTER_FILE} …")
    wb_master = Workbook()
    wb_master.remove(wb_master.active)

    # Summary sheet
    sum_ws = wb_master.create_sheet("SUMMARY")
    sum_ws.sheet_view.showGridLines = False
    sum_ws.merge_cells("A1:G1")
    c = sum_ws["A1"]
    c.value = "OpenFIGI Decode Summary — All Underlyings"
    c.font = hfont(13); c.fill = fill(NAVY); c.alignment = center()
    sum_ws.row_dimensions[1].height = 28

    sum_headers = ["Underlying","Total FIGIs","Decoded OK","Failed","Calls","Puts","Unique Expiries"]
    sum_widths  = [14, 14, 14, 10, 10, 10, 16]
    for ci, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        c = sum_ws.cell(row=3, column=ci, value=h)
        c.font = hfont(9); c.fill = fill(BLUE); c.alignment = center()
        sum_ws.column_dimensions[get_column_letter(ci)].width = w
    sum_ws.row_dimensions[3].height = 16

    totals = [0, 0, 0, 0, 0, 0]
    for ri, underlying in enumerate(sheet_figis):
        sub      = df_all[df_all["_underlying"] == underlying]
        ok       = int(sub["description"].notna().sum() if "description" in sub else 0)
        failed   = len(sub) - ok
        calls    = int((sub.get("opt_type", pd.Series()) == "Call").sum())
        puts     = int((sub.get("opt_type", pd.Series()) == "Put").sum())
        uniq_exp = sub["expiry"].nunique() if "expiry" in sub.columns else 0
        vals     = [underlying, len(sub), ok, failed, calls, puts, uniq_exp]
        bg       = WHITE if ri % 2 == 0 else LGREY
        for ci, val in enumerate(vals, 1):
            c = sum_ws.cell(row=ri+4, column=ci, value=val)
            c.font = dfont(9); c.fill = fill(bg); c.alignment = center(); c.border = thin()
        sum_ws.row_dimensions[ri+4].height = 14
        for i, v in enumerate([len(sub), ok, failed, calls, puts, uniq_exp]):
            totals[i] += v

    tr = len(sheet_figis) + 4
    for ci, val in enumerate(["TOTAL"] + totals, 1):
        c = sum_ws.cell(row=tr, column=ci, value=val)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = fill(NAVY); c.alignment = center(); c.border = thin()
    sum_ws.row_dimensions[tr].height = 16

    for underlying in sheet_figis:
        sub = df_all[df_all["_underlying"] == underlying][COLUMNS].copy()
        if sub.empty:
            continue
        df_to_sheet(wb_master.create_sheet(underlying[:31]), sub,
                    f"{underlying} Options — Decoded")

    wb_master.save(MASTER_FILE)
    log.info(f"✓ Master file saved: {MASTER_FILE}")

    # ── Final summary ─────────────────────────────────────────────────────────
    ok_mask = df_all["description"].notna() & (df_all["description"] != "")
    log.info("=" * 60)
    log.info(f"  Total FIGIs      : {len(df_all):,}")
    log.info(f"  Decoded OK       : {ok_mask.sum():,}  ({ok_mask.mean()*100:.1f}%)")
    log.info(f"  Failed / N/A     : {(~ok_mask).sum():,}")
    if "opt_type" in df_all.columns:
        log.info(f"  Calls            : {(df_all['opt_type']=='Call').sum():,}")
        log.info(f"  Puts             : {(df_all['opt_type']=='Put').sum():,}")
    if "expiry" in df_all.columns:
        log.info(f"  Unique expiries  : {df_all['expiry'].nunique():,}")
    if "strike" in df_all.columns:
        s = df_all["strike"].dropna()
        if len(s):
            log.info(f"  Strike range     : {s.min():.2f} – {s.max():.2f}")
    log.info(f"  Total run time   : {timedelta(seconds=int(time.time()-run_start))}")
    log.info("=" * 60)
    log.info(f"All output files in  {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
