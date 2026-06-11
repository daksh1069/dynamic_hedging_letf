#!/usr/bin/env python3
"""
Excel 5 — Historical Call Option Chain (Bloomberg formula template)

Builds the quarterly OPT_CHAIN template: one sheet per underlying, each with
26 columns (Q1 2020 -> Q2 2026), where each column is a single BDS formula
that spills every call-option BBG ID present in that underlying's chain on
the first day of that quarter:

    =BDS(security,"OPT_CHAIN","SINGLE_DATE_OVERRIDE","YYYYMMDD",
         "CHAIN_PUT_CALL_TYPE_OVRD","C")

Workflow: open the saved workbook on a Bloomberg Terminal, let the BDS
formulas refresh/spill, then "Paste Special -> Values" each sheet and save
as data/Excel5_OptionTickers_Final.xlsx -- the input to
excel_formula/scripts/01_process_excel5.py.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1]

TICKERS = [
    {"sheet": "MSTR", "bbg": "MSTR US Equity", "desc": "MicroStrategy (underlying of MSTU)",                       "flag": ""},
    {"sheet": "SMCI", "bbg": "SMCI US Equity", "desc": "Super Micro Computer (underlying of SMCX)",                "flag": ""},
    {"sheet": "TSLA", "bbg": "TSLA US Equity", "desc": "Tesla (underlying of TSLT)",                               "flag": ""},
    {"sheet": "COIN", "bbg": "COIN US Equity", "desc": "Coinbase (underlying of CONL)",                            "flag": ""},
    {"sheet": "PLTR", "bbg": "PLTR US Equity", "desc": "Palantir (underlying of PLTU)",                            "flag": ""},
    {"sheet": "MU",   "bbg": "MU US Equity",   "desc": "Micron (underlying of MUU)",                               "flag": ""},
    {"sheet": "XETH", "bbg": "XETH Curncy",    "desc": "Ethereum (underlying of ETHT) ⚠ Non-equity",          "flag": "⚠ CHECK"},
    {"sheet": "NG1",  "bbg": "NG1 Comdty",     "desc": "Natural Gas Futures (underlying of BOIL) ⚠ Non-equity", "flag": "⚠ CHECK"},
    {"sheet": "SOXX", "bbg": "SOXX US Equity", "desc": "Semiconductor ETF (underlying of SOXL)",                   "flag": ""},
    {"sheet": "AVGO", "bbg": "AVGO US Equity", "desc": "Broadcom (underlying of AVL)",                             "flag": ""},
    {"sheet": "NVO",  "bbg": "NVO US Equity",  "desc": "Novo Nordisk (underlying of NVOX)",                        "flag": ""},
    {"sheet": "NVDA", "bbg": "NVDA US Equity", "desc": "Nvidia (underlying of NVDL)",                              "flag": ""},
    {"sheet": "MSOS", "bbg": "MSOS US Equity", "desc": "Cannabis ETF (underlying of MSOX) ⚠ Thin liquidity",  "flag": "⚠ CHECK"},
]

NAVY = "1F3864"; BLUE = "2E75B6"; LBLUE = "BDD7EE"
ORANGE = "FCE4D6"; GREEN = "E2EFDA"; GRAY = "F2F2F2"
WHITE = "FFFFFF"; BLACK = "000000"; DKBLUE = "1F497D"

thin = Side(style="thin", color="CCCCCC")
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def st(cell, text, bg=NAVY, fg=WHITE, bold=True, size=11, italic=False, wrap=False, align="left"):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, italic=italic, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap, indent=1 if align == "left" else 0)


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return ws.cell(r1, c1)


def rh(ws, r, h): ws.row_dimensions[r].height = h
def cw(ws, c, w): ws.column_dimensions[get_column_letter(c)].width = w


def quarter_columns(start=(2020, 1), end=(2026, 2)):
    """Return [(label, 'YYYYMMDD'), ...] for each quarter start date, inclusive."""
    cols = []
    y, q = start
    while (y, q) <= end:
        month = (q - 1) * 3 + 1
        cols.append((f"Q{q} {y}\n({y}-{month:02d}-01)", f"{y}{month:02d}01"))
        q += 1
        if q > 4:
            q = 1
            y += 1
    return cols


def build():
    wb = Workbook(); wb.remove(wb.active)
    quarters = quarter_columns()
    n_cols = 1 + len(quarters)

    # ── INDEX ────────────────────────────────────────────────────────────────
    wi = wb.create_sheet("INDEX")
    wi.sheet_view.showGridLines = False

    c = merge(wi, 1, 1, 1, 5)
    st(c, "  Excel 5 — Historical Call Option Chain Directory", bg=NAVY, fg=WHITE, bold=True, size=14)
    rh(wi, 1, 28)

    c = merge(wi, 2, 1, 2, 5)
    st(c, '  Each sheet: quarterly BDS pulls of call option BBG IDs from Jan 2020 → Jun 2026  |  '
          'Formula: =BDS(security, "OPT_CHAIN", "SINGLE_DATE_OVERRIDE", date, "CHAIN_PUT_CALL_TYPE_OVRD", "C")',
       bg=BLUE, fg=WHITE, bold=False, size=9, italic=True, wrap=True)
    rh(wi, 2, 28)
    rh(wi, 3, 8)

    for ci, h in enumerate(["#", "Sheet", "Bloomberg Security", "Description", "Flag"], 1):
        c = wi.cell(4, ci); st(c, h, bg=BLUE, fg=WHITE, bold=True, size=9, align="center")
        c.border = box
    rh(wi, 4, 18)

    for idx, t in enumerate(TICKERS):
        row = idx + 5
        bg = ORANGE if t["flag"] else WHITE
        for ci, v in enumerate([idx + 1, t["sheet"], t["bbg"], t["desc"], t["flag"]], 1):
            cell = wi.cell(row, ci)
            cell.value = v
            cell.font = Font(name="Arial", size=10, bold=bool(t["flag"]) and ci == 5,
                              color="C00000" if (ci == 5 and t["flag"]) else BLACK)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = box
            cell.alignment = Alignment(horizontal="center" if ci in (1, 5) else "left", vertical="center", indent=1)
        rh(wi, row, 18)

    for ci, w in zip(range(1, 6), [4, 8, 18, 44, 12]):
        cw(wi, ci, w)

    # ── HOW_TO_USE ───────────────────────────────────────────────────────────
    how = wb.create_sheet("HOW_TO_USE")
    how.sheet_view.showGridLines = False

    c = merge(how, 1, 1, 1, 2)
    st(c, "  HOW TO USE EXCEL 5", bg=NAVY, fg=WHITE, bold=True, size=13)
    rh(how, 1, 28)
    rh(how, 2, 8)

    steps = [
        ("STEP 1", "Open this file on the Bloomberg Terminal. Ensure Bloomberg Excel Add-in is active "
                   "(DAPI <Go> on terminal to verify)."),
        ("STEP 2", "Navigate to any underlying sheet (e.g. MSTR, TSLA). Each sheet contains quarterly "
                   "BDS formulas from Q1 2020 → Q2 2026."),
        ("STEP 3", 'Each column = one quarter. The formula pulls every call option BBG ID that existed '
                   'on that date via:\n=BDS(security,"OPT_CHAIN","SINGLE_DATE_OVERRIDE","YYYYMMDD",'
                   '"CHAIN_PUT_CALL_TYPE_OVRD","C")'),
        ("STEP 4", "Bloomberg returns BBG Global IDs (e.g. BBG00P7YZLZ9 Equity). These are fully valid "
                   "Bloomberg tickers."),
        ("STEP 5", 'To decode any ID: =BDP("BBG00XXXXX Equity","TICKER") → human ticker\n'
                   '           =BDP("BBG00XXXXX Equity","OPT_EXPIRE_DT") → expiry date\n'
                   '           =BDP("BBG00XXXXX Equity","OPT_STRIKE_PX") → strike price'),
        ("STEP 6", "Use the BBG IDs directly in Excel 6 BDH formulas to pull historical Bid/Ask/IV/Greeks "
                   "for each contract."),
        ("⚠ FLAGS", "XETH, NG1, MSOS sheets are flagged — verify results manually. XETH/NG1 are "
                         "non-equity; MSOS has thin option liquidity."),
    ]
    for i, (label, text) in enumerate(steps):
        row = i + 3
        c = how.cell(row, 1)
        st(c, label, bg="C00000" if "FLAG" in label else NAVY, fg=WHITE, bold=True, size=9, align="center")
        c2 = how.cell(row, 2)
        st(c2, text, bg=GRAY, fg=BLACK, bold=False, size=9, wrap=True)
        rh(how, row, 60 if "\n" in text else 32)

    cw(how, 1, 12); cw(how, 2, 100)

    # ── Per-underlying sheets ────────────────────────────────────────────────
    for t in TICKERS:
        ws = wb.create_sheet(t["sheet"])
        ws.sheet_view.showGridLines = False
        if t["flag"]:
            ws.sheet_properties.tabColor = "FFC000"

        c = merge(ws, 1, 1, 1, n_cols)
        st(c, f"  {t['sheet']} — Historical Call Option Chain  |  {t['bbg']}  |  "
              f"OPT_CHAIN + SINGLE_DATE_OVERRIDE + CHAIN_PUT_CALL_TYPE_OVRD=C",
           bg=NAVY, fg=WHITE, bold=True, size=11)
        rh(ws, 1, 22)

        c = ws.cell(2, 1); st(c, "Contract #", bg=BLUE, fg=WHITE, bold=True, size=9, align="center", wrap=True)
        for qi, (label, _) in enumerate(quarters, 2):
            c = ws.cell(2, qi); st(c, label, bg=BLUE, fg=WHITE, bold=True, size=8, align="center", wrap=True)
        rh(ws, 2, 30)

        c = ws.cell(3, 1)
        st(c, "↓ BBG IDs spill below", bg=GRAY, fg="404040", bold=False, size=8, italic=True)
        c.border = box
        for qi, (_, date_str) in enumerate(quarters, 2):
            cell = ws.cell(3, qi)
            cell.value = (f'=BDS("{t["bbg"]}","OPT_CHAIN","SINGLE_DATE_OVERRIDE","{date_str}",'
                           f'"CHAIN_PUT_CALL_TYPE_OVRD","C")')
            cell.font = Font(name="Courier New", size=8, bold=True, color=DKBLUE)
            cell.fill = PatternFill("solid", fgColor=LBLUE)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = box
        rh(ws, 3, 18)

        ws.freeze_panes = ws.cell(4, 2)
        cw(ws, 1, 16)
        for qi in range(2, n_cols + 1):
            cw(ws, qi, 13)

        print(f"  {t['sheet']:6s}  ->  {len(quarters)} quarterly BDS columns  |  {t['bbg']}")

    out_path = OUT / "Excel5_OptionTickers_Final.xlsx"
    wb.save(out_path)
    print(f"\nSaved -> {out_path}")
    print("Next: open on Bloomberg Terminal, let BDS formulas spill, Paste Special -> Values")
    print("per sheet, then save as data/Excel5_OptionTickers_Final.xlsx for excel_formula/scripts/01.")


if __name__ == "__main__":
    build()
