#!/usr/bin/env python3
"""
Excel 3 -- Underlying Asset Data (Bloomberg formula template)

Builds one BDH formula sheet per underlying (13 names): daily Open/High/Low/
Close/Volume from 06/01/2020 -> today, plus Total Return Index and Market Cap
for the equity underlyings (OHLCV only for the crypto/commodity underlyings
ETH and NG1):

    =BDH(security,{"PX_OPEN","PX_HIGH","PX_LOW","PX_LAST",
         "TOT_RETURN_INDEX_GROSS_DVDS","PX_VOLUME","CUR_MKT_CAP"},
         "06/01/2020",TEXT(TODAY(),"MM/DD/YYYY"),
         "Dts=S","Dir=V","Per=CD","Days=W","CDR=5D")

Workflow: open the saved workbook on a Bloomberg Terminal, let the BDH
formulas spill, then "Paste Special -> Values" each sheet and save as
data/Excel3_Underlying_Data.xlsx -- used by
excel_formula/scripts/03_filter_contracts.py for the moneyness filter.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1]

UNDERLYINGS = [
    {"sheet": "MSTR", "bbg": "MSTR US Equity", "asset_class": "Equity",    "letf": "MSTU", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "SMCI", "bbg": "SMCI US Equity", "asset_class": "Equity",    "letf": "SMCX", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "TSLA", "bbg": "TSLA US Equity", "asset_class": "Equity",    "letf": "TSLT", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "COIN", "bbg": "COIN US Equity", "asset_class": "Equity",    "letf": "CONL", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "PLTR", "bbg": "PLTR US Equity", "asset_class": "Equity",    "letf": "PLTU", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "MU",   "bbg": "MU US Equity",   "asset_class": "Equity",    "letf": "MUU",  "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "ETH",  "bbg": "XETH Curncy",    "asset_class": "Crypto",    "letf": "ETHT", "options": "CME ETH Futures options",   "flag": "⚑",
     "flag_text": "  ⚠  FLAG: Crypto underlying. Options only via CME ETH futures — limited historical data depth."},
    {"sheet": "NG1",  "bbg": "NG1 Comdty",     "asset_class": "Commodity", "letf": "BOIL", "options": "UNG / NG Futures options",  "flag": "⚑",
     "flag_text": "  ⚠  FLAG: Commodity underlying. Use UNG US Equity or NG Comdty options — structure differs from equity options."},
    {"sheet": "SOXX", "bbg": "SOXX US Equity", "asset_class": "Equity",    "letf": "SOXL", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "AVGO", "bbg": "AVGO US Equity", "asset_class": "Equity",    "letf": "AVL",  "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "NVO",  "bbg": "NVO US Equity",  "asset_class": "Equity",    "letf": "NVOX", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "NVDA", "bbg": "NVDA US Equity", "asset_class": "Equity",    "letf": "NVDL", "options": "Standard listed options",   "flag": "",  "flag_text": ""},
    {"sheet": "MSOS", "bbg": "MSOS US Equity", "asset_class": "Equity",    "letf": "MSOX", "options": "Standard listed options",   "flag": "⚑",
     "flag_text": "  ⚠  FLAG: Cannabis ETF. Thin options market — expect wide bid-ask spreads. Treat option data with caution."},
]

START_DATE = "06/01/2020"

NAVY = "1F3864"; BLUE = "2E75B6"; LBLUE = "BDD7EE"
ORANGE = "FCE4D6"; GREEN = "E2EFDA"; GRAY = "F2F2F2"
WHITE = "FFFFFF"; BLACK = "000000"; YELLOW = "FFFF00"; INPUT_BLUE = "0000FF"

thin = Side(style="thin", color="CCCCCC")
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def st(cell, text, bg=NAVY, fg=WHITE, bold=True, size=11, italic=False, wrap=False, align="left"):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, italic=italic, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap, indent=1 if align == "left" else 0)


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return ws.cell(r1, c1)


def rh(ws, r, h): ws.row_dimensions[r].height = h
def cw(ws, c, w): ws.column_dimensions[get_column_letter(c)].width = w


def bdh_formula(bbg, equity):
    fields = ('"PX_OPEN","PX_HIGH","PX_LOW","PX_LAST","TOT_RETURN_INDEX_GROSS_DVDS","PX_VOLUME","CUR_MKT_CAP"'
              if equity else
              '"PX_OPEN","PX_HIGH","PX_LOW","PX_LAST","PX_VOLUME"')
    return (f'=BDH("{bbg}",{{{fields}}},"{START_DATE}",TEXT(TODAY(),"MM/DD/YYYY"),'
            f'"Dts=S","Dir=V","Per=CD","Days=W","CDR=5D")')


def build():
    wb = Workbook(); wb.remove(wb.active)

    # ── INDEX ────────────────────────────────────────────────────────────────
    wi = wb.create_sheet("INDEX")
    wi.sheet_view.showGridLines = False

    c = merge(wi, 1, 1, 1, 7)
    st(c, "  Excel 3 — Underlying Asset Data  |  Sheet Index", bg=NAVY, fg=WHITE, bold=True, size=13)
    rh(wi, 1, 26)

    c = merge(wi, 2, 1, 2, 7)
    st(c, "  One sheet per underlying. Equity: OHLCV + Total Return + Market Cap. Crypto/Commodity: OHLCV only.",
       bg=NAVY, fg=WHITE, bold=False, size=9)
    rh(wi, 2, 16)

    for ci, h in enumerate(["#", "Underlying", "Bloomberg Ticker", "Asset Class", "LETF Pair", "Options Available", "Flag"], 1):
        c = wi.cell(3, ci); st(c, h, bg=LBLUE, fg=BLACK, bold=True, size=9, align="center", wrap=True)
    rh(wi, 3, 30)

    for idx, u in enumerate(UNDERLYINGS):
        row = idx + 4
        for ci, v in enumerate([idx + 1, u["sheet"], u["bbg"], u["asset_class"], u["letf"], u["options"], u["flag"]], 1):
            cell = wi.cell(row, ci, v)
            cell.font = Font(name="Arial", size=10)
            if ci == 7 and u["flag"]:
                cell.fill = PatternFill("solid", fgColor=ORANGE)
        rh(wi, row, 18)

    for ci, w in zip(range(1, 8), [4, 10, 18, 12, 10, 26, 8]):
        cw(wi, ci, w)

    # ── Per-underlying sheets ────────────────────────────────────────────────
    for u in UNDERLYINGS:
        ws = wb.create_sheet(u["sheet"])
        ws.sheet_view.showGridLines = False
        equity = u["asset_class"] == "Equity"
        n_cols = 8 if equity else 6
        if u["flag"]:
            ws.sheet_properties.tabColor = "FFC000"

        c = merge(ws, 1, 1, 1, n_cols)
        st(c, f"  Underlying Data — {u['sheet']}  |  (LETF Pair: {u['letf']})", bg=NAVY, fg=WHITE, bold=True, size=13)
        rh(ws, 1, 26)

        c = merge(ws, 2, 1, 2, n_cols)
        st(c, f"  Bloomberg: {u['bbg']}  |  Asset Class: {u['asset_class']}  |  LETF: {u['letf']} US Equity",
           bg=NAVY, fg=WHITE, bold=False, size=9)
        rh(ws, 2, 16)

        rh(ws, 3, 7)
        r = 4

        if u["flag"]:
            c = merge(ws, r, 1, r, n_cols)
            st(c, u["flag_text"], bg=ORANGE, fg="C00000", bold=True, size=9, wrap=True)
            rh(ws, r, 32)
            r += 1
            rh(ws, r, 7)
            r += 1

        c = merge(ws, r, 1, r, n_cols)
        st(c, "  Parameters", bg=BLUE, fg=WHITE, bold=True, size=10)
        rh(ws, r, 20)
        r += 1

        c = ws.cell(r, 1); st(c, "  Bloomberg Ticker:", bg=WHITE, fg=BLACK, bold=True, size=10)
        c2 = ws.cell(r, 2); st(c2, u["bbg"], bg=YELLOW, fg=INPUT_BLUE, bold=False, size=10)
        rh(ws, r, 18)
        r += 1

        c = ws.cell(r, 1); st(c, "  Start Date:", bg=WHITE, fg=BLACK, bold=True, size=10)
        c2 = ws.cell(r, 2); st(c2, START_DATE, bg=YELLOW, fg=INPUT_BLUE, bold=False, size=10)
        c3 = merge(ws, r, 4, r, n_cols)
        st(c3, "  Bloomberg returns max available history if asset inception is after start date.",
           bg=WHITE, fg="404040", bold=False, size=9, italic=True)
        rh(ws, r, 18)
        r += 1

        c = ws.cell(r, 1); st(c, "  End Date:", bg=WHITE, fg=BLACK, bold=True, size=10)
        c2 = ws.cell(r, 2); st(c2, '=TEXT(TODAY(),"MM/DD/YYYY")', bg=YELLOW, fg=INPUT_BLUE, bold=False, size=10)
        rh(ws, r, 18)
        r += 1

        rh(ws, r, 7)
        r += 1

        c = merge(ws, r, 1, r, n_cols)
        st(c, "  Historical Daily Data — BDH", bg=BLUE, fg=WHITE, bold=True, size=10)
        rh(ws, r, 20)
        r += 1

        src = (f"  Source: Bloomberg Terminal | {u['bbg']}  |  OHLCV + Total Return Index + Market Cap"
               if equity else
               f"  Source: Bloomberg Terminal | {u['bbg']}  |  OHLCV only (no market cap for {u['asset_class']})")
        c = merge(ws, r, 1, r, n_cols)
        st(c, src, bg=GREEN, fg="404040", bold=False, size=9, wrap=True)
        rh(ws, r, 22)
        r += 1

        headers = (["Date", "Open", "High", "Low", "Close", "Total Return Idx", "Volume", "Mkt Cap ($M)"]
                   if equity else
                   ["Date", "Open", "High", "Low", "Close", "Volume"])
        for ci, h in enumerate(headers, 1):
            c = ws.cell(r, ci); st(c, h, bg=LBLUE, fg=BLACK, bold=True, size=9, align="center", wrap=True)
        rh(ws, r, 30)
        r += 1

        cell = ws.cell(r, 1)
        cell.value = bdh_formula(u["bbg"], equity)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top")
        rh(ws, r, 16)

        cw(ws, 1, 13)
        for ci in range(2, n_cols + 1):
            cw(ws, ci, 15)

        print(f"  {u['sheet']:6s}  ->  {u['asset_class']:9s}  |  {u['bbg']}")

    out_path = OUT / "Excel3_Underlying_Data.xlsx"
    wb.save(out_path)
    print(f"\nSaved -> {out_path}")
    print("Next: open on Bloomberg Terminal, let BDH formulas spill, Paste Special -> Values")
    print("per sheet, then save as data/Excel3_Underlying_Data.xlsx for excel_formula/scripts/03.")


if __name__ == "__main__":
    build()
