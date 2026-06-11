"""
Sanity-check data/<TICKER>_calls_PXLAST_full_filled.xlsx.

For each Batch_NN sheet, walk each [Date | PX_LAST | PX_VOLUME | blank] block
(one per security) and check:
  - the BDH formula actually evaluated (no #NAME?/#N/A/error strings left)
  - how many data rows it returned, and the date range covered
  - flags any block with zero rows or error values

Usage: python excel_formula/scripts/06_check_bdh_full.py TICKER
"""
import datetime
import sys
from pathlib import Path

import openpyxl

if len(sys.argv) < 2:
    sys.exit("Usage: python excel_formula/scripts/06_check_bdh_full.py TICKER")
TICKER = sys.argv[1]

ROOT = Path(__file__).resolve().parents[2]
PATH = ROOT / "data" / f"{TICKER}_calls_PXLAST_full_filled.xlsx"

BLOCK_WIDTH = 4
ERROR_STRINGS = {"#NAME?", "#N/A", "#REF!", "#VALUE!", "#NULL!", "#DIV/0!", "#NUM!"}


def main():
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)

    total_securities = 0
    total_rows = 0
    empty_blocks = []
    error_blocks = []
    min_date = None
    max_date = None

    for name in wb.sheetnames:
        if not name.startswith("Batch"):
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        max_col = max(len(r) for r in rows)
        n_blocks = (max_col + BLOCK_WIDTH - 1) // BLOCK_WIDTH

        for b in range(n_blocks):
            c0 = b * BLOCK_WIDTH  # 0-indexed: Date col
            label = rows[0][c0] if c0 < len(rows[0]) else None
            if label is None:
                continue  # padding block beyond actual securities
            total_securities += 1

            n_data = 0
            block_has_error = False
            for r in rows[2:]:  # row index 2 = Excel row 3 (0-indexed)
                if c0 >= len(r):
                    continue
                date_val = r[c0]
                px_val = r[c0 + 1] if c0 + 1 < len(r) else None
                if date_val is None and px_val is None:
                    continue
                if isinstance(date_val, str) and date_val.strip() in ERROR_STRINGS:
                    block_has_error = True
                    continue
                if isinstance(px_val, str) and px_val.strip() in ERROR_STRINGS:
                    block_has_error = True
                    continue
                if isinstance(date_val, datetime.datetime):
                    if min_date is None or date_val < min_date:
                        min_date = date_val
                    if max_date is None or date_val > max_date:
                        max_date = date_val
                n_data += 1

            total_rows += n_data
            if n_data == 0:
                empty_blocks.append((name, label))
            if block_has_error:
                error_blocks.append((name, label))

    print(f"Securities found: {total_securities:,} (expect 3,797)")
    print(f"Total data rows (date x security): {total_rows:,}")
    print(f"Date range: {min_date.date() if min_date else None} -> {max_date.date() if max_date else None}")
    print(f"Empty blocks (0 rows): {len(empty_blocks)}")
    for s in empty_blocks[:20]:
        print(f"  {s}")
    print(f"Blocks with leftover error values: {len(error_blocks)}")
    for s in error_blocks[:20]:
        print(f"  {s}")


if __name__ == "__main__":
    main()
