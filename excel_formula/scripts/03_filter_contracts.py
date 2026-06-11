"""
Stage 0c: read all data/decoded/*_decoded.xlsx files
(output of 02_decode_openfigi.py), filter to the call options relevant for
the backtest, and write into data/filtered/:
  - <TICKER>_calls_filtered.xlsx  (one per underlying)
  - ALL_calls_filtered.xlsx        (master summary)
  - Excel6_BDH_Ready.xlsx          (Bloomberg BDH formulas -- superseded,
    only PX_LAST is needed; kept for reference)

Filter criteria applied in order:
  1. Calls only                (drops ~50%)
  2. Valid expiry 2020-2026    (drops bad/NaN rows)
  3. DTE window 15-180 days   (drops deep LEAPS and weekly junk)
     at ANY quarterly obs date (Jan/Apr/Jul/Oct 1 each year)
  4. Moneyness 80%-130%       (needs underlying price history --
     defaults to data/Excel3_Underlying_Data.xlsx)
     at the quarterly obs date closest to each contract's midlife

Usage
-----
  python excel_formula/scripts/03_filter_contracts.py

  # Without the moneyness filter (faster, still cuts 60-70%)
  python excel_formula/scripts/03_filter_contracts.py --excel3 ""

Output: data/filtered/TSLA_calls_filtered.xlsx (our
3,797-contract universe) + the same for every other team ticker, plus
ALL_calls_filtered.xlsx, Excel6_BDH_Ready.xlsx, filter_summary.txt.
"""

import os
import sys
import glob
import argparse
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"

DECODED_DIR   = DATA_DIR / "decoded"
OUTPUT_DIR    = DATA_DIR / "filtered"
EXCEL3_FILE   = DATA_DIR / "Excel3_Underlying_Data.xlsx"
SKIP_FILES    = {"ALL_decoded.xlsx", "_checkpoint.json"}

MIN_DTE       = 15      # minimum days to expiry at observation date
MAX_DTE       = 180     # maximum days to expiry at observation date
MIN_MONEYNESS = 0.80    # 80% of underlying price
MAX_MONEYNESS = 1.30    # 130% of underlying price

# Quarterly observation dates: Jan 1, Apr 1, Jul 1, Oct 1 from 2020 to 2026
OBS_DATES = pd.date_range(start="2020-01-01", end="2026-10-01", freq="QS")

# Bloomberg fields to pull in Excel 6
BDH_FIELDS = "PX_BID,PX_ASK,IVOL_MID,DELTA_MID,GAMMA_MID,THETA_MID,VEGA_MID,OPT_OPEN_INT"
BDH_START  = "20200101"
BDH_END    = "20261231"

# ── UNDERLYING → Bloomberg security ticker mapping (for Excel 3 lookup) ──────
UNDERLYING_BBG = {
    "MSTR": "MSTR US Equity",
    "SMCI": "SMCI US Equity",
    "TSLA": "TSLA US Equity",
    "COIN": "COIN US Equity",
    "PLTR": "PLTR US Equity",
    "MU":   "MU US Equity",
    "XETH": "XETH Curncy",
    "NG1":  "NG1 Comdty",
    "SOXX": "SOXX US Equity",
    "AVGO": "AVGO US Equity",
    "NVO":  "NVO US Equity",
    "NVDA": "NVDA US Equity",
    "MSOS": "MSOS US Equity",
}

# ── COLORS ────────────────────────────────────────────────────────────────────
NAVY  = "1F3864"
BLUE  = "2E75B6"
GREEN = "E2EFDA"
AMBER = "FFF2CC"
RED   = "FFDCD8"
WHITE = "FFFFFF"
GREY  = "F2F2F2"

def fill(c): return PatternFill("solid", fgColor=c)
def hf(s=9): return Font(name="Arial", bold=True,  size=s, color="FFFFFF")
def df(s=9): return Font(name="Arial", bold=False, size=s, color="000000")
def mf(s=8): return Font(name="Courier New", size=s, color="000080")
def ctr():   return Alignment(horizontal="center", vertical="center")
def lft():   return Alignment(horizontal="left",   vertical="center")
def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ── STEP 1: READ ALL DECODED FILES ────────────────────────────────────────────
def read_decoded_files(decoded_dir: str) -> "dict[str, pd.DataFrame]":
    pattern = os.path.join(decoded_dir, "*_decoded.xlsx")
    files   = sorted(glob.glob(pattern))
    if not files:
        print(f"ERROR: No *_decoded.xlsx files found in '{decoded_dir}/'")
        print("Make sure you run this script from the same folder as the decoded/ directory.")
        sys.exit(1)

    all_dfs = {}
    for fpath in files:
        fname = os.path.basename(fpath)
        if fname in SKIP_FILES:
            continue
        ticker = fname.replace("_decoded.xlsx", "")
        try:
            df = pd.read_excel(fpath, header=2)   # row 3 is the header (0-indexed row 2)
            df.columns = [c.strip() for c in df.columns]
            # Rename columns to standard names
            rename = {
                "FIGI":                 "figi",
                "Raw BBG ID":           "raw_id",
                "Security Description": "description",
                "Ticker":               "ticker",
                "Underlying":           "underlying",
                "Exchange":             "exchange",
                "Expiry Date":          "expiry",
                "Strike":               "strike",
                "Opt Type":             "opt_type",
                "Security Type":        "security_type",
                "Market Sector":        "market_sector",
            }
            df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
            df["_source_ticker"] = ticker
            all_dfs[ticker] = df
            print(f"  Read {ticker:8s}: {len(df):>6,} rows")
        except Exception as e:
            print(f"  WARNING: Could not read {fname}: {e}")

    return all_dfs


# ── STEP 2: LOAD UNDERLYING PRICES FROM EXCEL 3 (optional) ───────────────────
def load_underlying_prices(excel3_path: str) -> "dict[str, pd.Series]":
    """Returns {ticker: pd.Series(index=date, values=close_price)}"""
    if not excel3_path or not os.path.exists(excel3_path):
        return {}

    print(f"\nLoading underlying prices from {excel3_path} ...")
    prices = {}
    try:
        wb = load_workbook(excel3_path, data_only=True)
    except Exception as e:
        print(f"  WARNING: Could not open Excel 3: {e}")
        return prices

    for sheet in wb.sheetnames:
        if sheet in {"INDEX", "HOW_TO_USE"}:
            continue
        try:
            ws   = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))

            # Find the exact data header row — first column must be literally "Date"
            # (avoids matching parameter rows like "  Start Date:")
            header_row = None
            for i, row in enumerate(rows):
                if row[0] is not None and str(row[0]).strip().lower() == "date":
                    header_row = i
                    break
            if header_row is None:
                continue

            headers   = [str(c).strip() if c else "" for c in rows[header_row]]
            date_col  = next((i for i, h in enumerate(headers) if h.lower() == "date"), None)
            price_col = next((i for i, h in enumerate(headers)
                              if any(x in h.upper() for x in ["PX_LAST","CLOSE","LAST"])), None)
            if date_col is None or price_col is None:
                continue

            pairs = []   # collect (date, price) atomically — never append one without the other
            for row in rows[header_row + 1:]:
                d = row[date_col]
                p = row[price_col]
                if d is None or p is None:
                    continue
                try:
                    pairs.append((pd.to_datetime(d), float(p)))
                except Exception:
                    pass

            if pairs:
                dates, pxs = zip(*pairs)
                s = pd.Series(list(pxs), index=pd.DatetimeIndex(dates)).sort_index()
                prices[sheet] = s
                print(f"  {sheet:8s}: {len(s):,} obs  "
                      f"{s.index[0].date()} → {s.index[-1].date()}")
        except Exception as e:
            print(f"  WARNING: Could not load sheet '{sheet}': {e}")

    return prices


# ── STEP 3: FILTER CONTRACTS ──────────────────────────────────────────────────
def filter_calls(df: pd.DataFrame, ticker: str,
                 prices: dict[str, pd.Series]) -> pd.DataFrame:

    # 3a. Calls only
    df = df[df["opt_type"].str.strip().str.lower() == "call"].copy()
    if df.empty:
        return df

    # 3b. Parse expiry dates
    df["expiry"] = pd.to_datetime(df["expiry"], errors="coerce")
    df = df.dropna(subset=["expiry"])
    df = df[(df["expiry"] >= "2020-01-01") & (df["expiry"] <= "2026-12-31")]
    if df.empty:
        return df

    # 3c. DTE window filter — keep contracts where at least one obs_date
    #     falls in [expiry - MAX_DTE, expiry - MIN_DTE]
    expiries  = df["expiry"].values                    # numpy datetime64 array
    dte_keep  = np.zeros(len(df), dtype=bool)

    for obs in OBS_DATES:
        obs_np    = np.datetime64(obs, "D")
        dte_vals  = (expiries.astype("datetime64[D]") - obs_np).astype(int)
        dte_keep |= (dte_vals >= MIN_DTE) & (dte_vals <= MAX_DTE)

    df = df[dte_keep].copy()
    if df.empty:
        return df

    # 3d. Optional moneyness filter — needs underlying price history
    price_series = prices.get(ticker)
    if price_series is not None and not price_series.empty:
        # For each contract, find the obs_date closest to midlife and get price
        df["expiry_dt"] = df["expiry"]

        def moneyness_ok(row):
            exp     = row["expiry_dt"]
            strike  = row["strike"]
            if pd.isna(strike):
                return True  # can't filter, keep

            # obs dates that bracket this contract (within DTE window)
            valid_obs = [o for o in OBS_DATES
                         if MIN_DTE <= (exp - o).days <= MAX_DTE]
            if not valid_obs:
                return True

            # For each valid obs date, check price and moneyness
            for obs in valid_obs:
                # Get nearest available price to obs date
                idx = price_series.index.searchsorted(obs)
                if idx >= len(price_series):
                    idx = len(price_series) - 1
                price = price_series.iloc[idx]
                if price > 0:
                    m = strike / price
                    if MIN_MONEYNESS <= m <= MAX_MONEYNESS:
                        return True
            return False

        mask = df.apply(moneyness_ok, axis=1)
        df   = df[mask].copy()

    return df


# ── STEP 4: BUILD EXCEL 6 BDH SHEET ──────────────────────────────────────────
def build_excel6(all_filtered: dict[str, pd.DataFrame]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # INDEX / HOW TO USE
    idx = wb.create_sheet("INDEX")
    idx.sheet_view.showGridLines = False

    idx.merge_cells("A1:G1")
    c = idx["A1"]
    c.value = "Excel 6 — BDH Historical Options Data  |  Open on Bloomberg Terminal"
    c.font = hf(13); c.fill = fill(NAVY); c.alignment = ctr()
    idx.row_dimensions[1].height = 28

    idx.merge_cells("A2:G2")
    c = idx["A2"]
    c.value = (f"Fields: {BDH_FIELDS}  |  "
               f"Period: {BDH_START} → {BDH_END}  |  "
               f"One sheet per underlying")
    c.font = Font(name="Arial", size=8, italic=True, color="FFFFFF")
    c.fill = fill(BLUE); c.alignment = lft()
    idx.row_dimensions[2].height = 14

    for col, h in enumerate(["#","Sheet","Contracts","DTE 15-180","Calls Only","→ Bloomberg"], 1):
        c = idx.cell(row=4, column=col, value=h)
        c.font = hf(9); c.fill = fill(BLUE); c.alignment = ctr()

    for i, (ticker, cdf) in enumerate(all_filtered.items()):
        row = i + 5
        bg  = WHITE if i % 2 == 0 else GREY
        for col, val in enumerate([i+1, ticker, len(cdf), "✓", "✓", "Open & Refresh"], 1):
            c = idx.cell(row=row, column=col, value=val)
            c.font = df_(9) if col > 1 else hf(9)
            c.fill = fill(bg); c.alignment = ctr(); c.border = thin()
        idx.row_dimensions[row].height = 14

    for col, w in zip("ABCDEFG", [6, 12, 14, 12, 12, 22]):
        idx.column_dimensions[col].width = w

    # One sheet per underlying
    COL_HEADERS = ["BBG ID (Security)", "Description", "Expiry", "Strike",
                   "BDH — Bid", "BDH — Ask", "BDH — IV", "BDH — Delta",
                   "BDH — Gamma", "BDH — Theta", "BDH — Vega", "BDH — OI"]
    COL_WIDTHS  = [28, 28, 14, 10, 28, 28, 28, 28, 28, 28, 28, 28]

    for ticker, cdf in all_filtered.items():
        if cdf.empty:
            continue

        ws = wb.create_sheet(ticker[:31])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5"

        # Title
        last = get_column_letter(len(COL_HEADERS))
        ws.merge_cells(f"A1:{last}1")
        c = ws["A1"]
        c.value = (f"{ticker} — {len(cdf):,} Call Options  |  "
                   f"BDH: {BDH_FIELDS.replace(',', ' | ')}")
        c.font = hf(10); c.fill = fill(NAVY); c.alignment = lft()
        ws.row_dimensions[1].height = 22

        ws.merge_cells(f"A2:{last}2")
        c = ws["A2"]
        c.value = ("Refresh on Bloomberg Terminal → select all → Paste Special Values → Save  |  "
                   f"Period: {BDH_START} → {BDH_END}")
        c.font = Font(name="Arial", size=8, italic=True, color="FFFFFF")
        c.fill = fill(BLUE); c.alignment = lft()
        ws.row_dimensions[2].height = 14

        ws.merge_cells(f"A3:{last}3")
        c = ws["A3"]
        c.value = ("NOTE: Each BDH formula returns a TIME SERIES (multiple rows). "
                   "Bloomberg will spill rows downward. Leave many blank rows below each formula row. "
                   "Better: use the Data tab → Bloomberg → Import for bulk pulls.")
        c.font = Font(name="Arial", size=8, italic=True, color="CC0000")
        c.fill = fill(AMBER); c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 30

        # Column headers
        for ci, (h, w) in enumerate(zip(COL_HEADERS, COL_WIDTHS), 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font = hf(9); c.fill = fill(BLUE); c.alignment = ctr()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[4].height = 16

        # Data rows — one per contract
        df_sorted = cdf.sort_values(["expiry", "strike"]).reset_index(drop=True)
        for ri, row_data in df_sorted.iterrows():
            r   = ri + 5
            bg  = WHITE if ri % 2 == 0 else GREY
            sec = str(row_data.get("raw_id", "")).strip()
            desc = str(row_data.get("description", "")).strip()
            exp  = row_data.get("expiry")
            exp_str = exp.strftime("%Y-%m-%d") if pd.notna(exp) else ""
            strike_val = row_data.get("strike", "")

            # Individual field BDH formulas
            field_list = BDH_FIELDS.split(",")
            cells_data = [sec, desc, exp_str, strike_val]
            for field in field_list:
                formula = f'=BDH("{sec}","{field}","{BDH_START}","{BDH_END}")'
                cells_data.append(formula)

            for ci, val in enumerate(cells_data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                if ci == 1:
                    c.font = mf(8); c.fill = fill(GREEN)
                elif ci <= 4:
                    c.font = df(9); c.fill = fill(bg)
                else:
                    c.font = Font(name="Arial", size=8, color="000080")
                    c.fill = fill(AMBER)
                c.alignment = ctr() if ci in [3,4] else lft()
                c.border = thin()
            ws.row_dimensions[r].height = 13

    return wb


def df_(s=9):
    return Font(name="Arial", bold=False, size=s, color="000000")


# ── STEP 5: WRITE SUMMARY ─────────────────────────────────────────────────────
def write_summary(all_raw: dict, all_filtered: dict,
                  excel3_used: bool, output_dir: str):
    lines = [
        "=" * 65,
        "FILTER SUMMARY",
        f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Moneyness filter: {'YES (80%-130%)' if excel3_used else 'NO (Excel3 not provided)'}",
        f"DTE filter: {MIN_DTE}-{MAX_DTE} days",
        "=" * 65,
        f"{'Underlying':<12} {'Raw':>8} {'Calls':>8} {'Filtered':>10} {'Cut %':>8}",
        "-" * 65,
    ]
    total_raw = total_calls = total_filt = 0
    for ticker in sorted(all_raw.keys()):
        raw   = len(all_raw[ticker])
        calls = int((all_raw[ticker].get("opt_type","").str.lower() == "call").sum())
        filt  = len(all_filtered.get(ticker, pd.DataFrame()))
        cut   = (1 - filt / calls) * 100 if calls > 0 else 0
        lines.append(f"{ticker:<12} {raw:>8,} {calls:>8,} {filt:>10,} {cut:>7.1f}%")
        total_raw += raw; total_calls += calls; total_filt += filt
    cut_total = (1 - total_filt / total_calls) * 100 if total_calls > 0 else 0
    lines += [
        "-" * 65,
        f"{'TOTAL':<12} {total_raw:>8,} {total_calls:>8,} {total_filt:>10,} {cut_total:>7.1f}%",
        "=" * 65,
        "",
        f"Output files: {output_dir}/",
        "  → *_calls_filtered.xlsx  (one per underlying)",
        "  → ALL_calls_filtered.xlsx",
        "  → Excel6_BDH_Ready.xlsx  ← take this to Bloomberg terminal",
    ]
    summary_text = "\n".join(lines)
    print("\n" + summary_text)
    path = os.path.join(output_dir, "filter_summary.txt")
    with open(path, "w") as f:
        f.write(summary_text)


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--decoded-dir", default=DECODED_DIR,
                        help=f"Folder containing *_decoded.xlsx files (default: {DECODED_DIR})")
    parser.add_argument("--excel3", default=EXCEL3_FILE,
                        help=f"Path to Excel3_Underlying_Data.xlsx for moneyness filter, "
                             f"or '' to skip it (default: {EXCEL3_FILE})")
    parser.add_argument("--output-dir", default=OUTPUT_DIR,
                        help=f"Output folder (default: {OUTPUT_DIR})")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    print("=" * 65)
    print("Contract Filter — starting")
    print(f"  Decoded dir : {args.decoded_dir}/")
    print(f"  Excel3      : {args.excel3 or 'Not provided — skipping moneyness filter'}")
    print(f"  Output dir  : {args.output_dir}/")
    print(f"  DTE range   : {MIN_DTE}–{MAX_DTE} days")
    print(f"  Moneyness   : {MIN_MONEYNESS*100:.0f}%–{MAX_MONEYNESS*100:.0f}% (if Excel3 provided)")
    print("=" * 65)

    # Read inputs
    print(f"\nReading decoded files from {args.decoded_dir}/ ...")
    all_raw = read_decoded_files(args.decoded_dir)

    prices = load_underlying_prices(args.excel3)
    excel3_used = bool(prices)

    # Filter each underlying
    print("\nFiltering contracts ...")
    all_filtered = {}
    all_filtered_rows = []

    for ticker, df in all_raw.items():
        filtered = filter_calls(df, ticker, prices)
        all_filtered[ticker] = filtered
        all_filtered_rows.append(filtered)

        calls_raw = int((df.get("opt_type", pd.Series()).str.lower() == "call").sum())
        print(f"  {ticker:8s}: {calls_raw:>5,} calls → {len(filtered):>5,} filtered  "
              f"({(1-len(filtered)/calls_raw)*100:.0f}% cut)" if calls_raw > 0
              else f"  {ticker:8s}: 0 calls")

    # Write per-underlying files
    print(f"\nWriting filtered files to {args.output_dir}/ ...")
    for ticker, df in all_filtered.items():
        if df.empty:
            continue
        path = os.path.join(args.output_dir, f"{ticker}_calls_filtered.xlsx")
        wb = Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet(ticker[:31])
        # Write header
        cols = ["raw_id","description","expiry","strike","figi","underlying","exchange","security_type"]
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=h.upper())
            c.font = hf(9); c.fill = fill(BLUE); c.alignment = ctr()
        for ri, (_, row_data) in enumerate(df.iterrows(), 2):
            for ci, col in enumerate(cols, 1):
                val = row_data.get(col, "")
                if isinstance(val, pd.Timestamp):
                    val = val.strftime("%Y-%m-%d")
                ws.cell(row=ri, column=ci, value=val).font = df_(9)
        wb.save(path)
        print(f"  ✓ {path}  ({len(df):,} contracts)")

    # Write master filtered file
    if all_filtered_rows:
        master_df = pd.concat([df for df in all_filtered_rows if not df.empty], ignore_index=True)
        master_path = os.path.join(args.output_dir, "ALL_calls_filtered.xlsx")
        master_df.to_excel(master_path, index=False)
        print(f"  ✓ {master_path}  ({len(master_df):,} total contracts)")

    # Build Excel 6
    print(f"\nBuilding Excel6_BDH_Ready.xlsx ...")
    wb6 = build_excel6(all_filtered)
    excel6_path = os.path.join(args.output_dir, "Excel6_BDH_Ready.xlsx")
    wb6.save(excel6_path)
    print(f"  ✓ {excel6_path}  ← take this to Bloomberg terminal")

    # Summary
    write_summary(all_raw, all_filtered, excel3_used, args.output_dir)


if __name__ == "__main__":
    main()
