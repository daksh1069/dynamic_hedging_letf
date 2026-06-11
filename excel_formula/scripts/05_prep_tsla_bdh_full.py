"""
Build the FULL Bloomberg BDH prep workbook for all 3,797 TSLA call options'
closing prices (PX_LAST) and volume (PX_VOLUME), 2020-01-01 -> 2026-06-10.

Validated approach (see TSLA_calls_PXLAST_prep_filled.xlsx, Sample_BDH sheet):
horizontal blocks of [Date | PX_LAST | PX_VOLUME | blank] per security, one
=BDH() formula per block, spilling downward. This avoids the row-collision
problem of one-row-per-security templates.

Batched into sheets of BATCH_SIZE securities (=BATCH_SIZE*4 columns) so the
user can refresh + "Paste Special -> Values" one sheet at a time on a
Bloomberg Terminal, checkpointing as they go.

Output: data/TSLA_calls_PXLAST_full.xlsx

No Databento/Bloomberg API calls -- pure local file generation, free.
"""
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "data"

START = "1/1/2020"
END = "6/10/2026"
BATCH_SIZE = 200
BLOCK_WIDTH = 4  # Date | PX_LAST | PX_VOLUME | blank

INSTRUCTIONS = [
    "TSLA call option closing prices -- FULL Bloomberg BDH workbook",
    "",
    "3,797 contracts, batched into sheets of "
    f"{BATCH_SIZE} securities ({BATCH_SIZE * BLOCK_WIDTH} columns each).",
    "Each security gets a [Date | PX_LAST | PX_VOLUME | blank] block with one",
    "=BDH() formula in row 3 that spills downward.",
    "",
    "Per-batch workflow (do ONE sheet at a time):",
    "  1. Open the 'Batch_NN' sheet. Formulas evaluate automatically if",
    "     connected to a live Bloomberg Terminal (replace #NAME? with data).",
    "  2. Wait for all blocks to finish spilling (no more #N/A Requesting...).",
    "  3. Select the whole sheet -> Copy -> Paste Special -> Values, to",
    "     freeze the results (so they don't re-query / break on next open).",
    "  4. Save. Move to the next Batch_NN sheet.",
    "",
    "If a batch is slow/unreliable, reduce BATCH_SIZE in",
    "excel_formula/scripts/05_prep_tsla_bdh_full.py and regenerate -- this is",
    "free and instant (no API cost, pure local formula generation).",
    "",
    "The 'Securities' sheet lists every (RAW_ID, expiry, strike) in the same",
    "order the batches were built, for reference when merging results back",
    "with the Databento-resolved contract mapping (kept outside this repo).",
]


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "TSLA_calls_PXLAST_full.xlsx"

    filt = pd.read_excel(
        ROOT / "data" / "filtered" / "TSLA_calls_filtered.xlsx",
        sheet_name="TSLA",
    )
    sec = filt[["RAW_ID", "DESCRIPTION", "EXPIRY", "STRIKE", "FIGI"]].sort_values(
        ["EXPIRY", "STRIKE"]
    ).reset_index(drop=True)
    print(f"Securities: {len(sec):,} rows")

    n_batches = -(-len(sec) // BATCH_SIZE)  # ceil
    print(f"Batches: {n_batches} x up to {BATCH_SIZE} securities "
          f"({BATCH_SIZE * BLOCK_WIDTH} cols/sheet)")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sec.to_excel(writer, sheet_name="Securities", index=False)
        pd.DataFrame({"Instructions": INSTRUCTIONS}).to_excel(
            writer, sheet_name="Instructions", index=False, header=False
        )
        for b in range(n_batches):
            pd.DataFrame().to_excel(writer, sheet_name=f"Batch_{b+1:02d}", index=False)

    wb = openpyxl.load_workbook(out_path)
    wb["Instructions"].column_dimensions["A"].width = 90

    sec_ws = wb["Securities"]
    for col_letter, width in zip("ABCDE", [22, 28, 12, 10, 14]):
        sec_ws.column_dimensions[col_letter].width = width

    for b in range(n_batches):
        ws = wb[f"Batch_{b+1:02d}"]
        chunk = sec.iloc[b * BATCH_SIZE: (b + 1) * BATCH_SIZE]
        for i, (_, row) in enumerate(chunk.iterrows()):
            col0 = i * BLOCK_WIDTH + 1  # 1-indexed
            raw_id = row["RAW_ID"]
            ws.cell(row=1, column=col0, value=f"{raw_id} ({row['DESCRIPTION']})")
            ws.cell(row=2, column=col0, value="Date")
            ws.cell(row=2, column=col0 + 1, value="PX_LAST")
            ws.cell(row=2, column=col0 + 2, value="PX_VOLUME")
            ws.cell(
                row=3,
                column=col0,
                value=f'=BDH("{raw_id}","PX_LAST,PX_VOLUME","{START}","{END}")',
            )

    wb.save(out_path)
    print(f"Saved -> {out_path}")


if __name__ == "__main__":
    main()
