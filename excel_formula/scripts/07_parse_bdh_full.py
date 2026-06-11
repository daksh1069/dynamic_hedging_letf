"""
Parse data/<TICKER>_calls_PXLAST_full_filled.xlsx (the
Bloomberg BDH pull: PX_LAST + PX_VOLUME for all filtered <TICKER> calls,
2020-01-02 -> 2026-06-10) into a clean long-format table.

Output: data/processed/<TICKER>_calls_close.parquet
  columns: raw_id, figi, expiry, strike, date, px_last, px_volume

Usage: python excel_formula/scripts/07_parse_bdh_full.py TICKER
"""
import datetime
import sys
from pathlib import Path

import openpyxl
import pandas as pd

if len(sys.argv) < 2:
    sys.exit("Usage: python excel_formula/scripts/07_parse_bdh_full.py TICKER")
TICKER = sys.argv[1]

ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"
PATH = DATA_DIR / f"{TICKER}_calls_PXLAST_full_filled.xlsx"

BLOCK_WIDTH = 4


def main():
    sec = pd.read_excel(PATH, sheet_name="Securities")
    sec_by_raw_id = sec.set_index("RAW_ID")[["FIGI", "EXPIRY", "STRIKE"]]

    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)

    rows = []
    for name in wb.sheetnames:
        if not name.startswith("Batch"):
            continue
        ws = wb[name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        max_col = max(len(r) for r in sheet_rows)
        n_blocks = (max_col + BLOCK_WIDTH - 1) // BLOCK_WIDTH

        for b in range(n_blocks):
            c0 = b * BLOCK_WIDTH
            label = sheet_rows[0][c0] if c0 < len(sheet_rows[0]) else None
            if label is None:
                continue
            raw_id = label.split(" (")[0].strip()

            for r in sheet_rows[2:]:
                if c0 + 1 >= len(r):
                    continue
                date_val, px_last, px_vol = r[c0], r[c0 + 1], r[c0 + 2] if c0 + 2 < len(r) else None
                if not isinstance(date_val, datetime.datetime):
                    continue  # skip stray "#N/A N/A" end-of-data markers
                rows.append((raw_id, date_val, px_last, px_vol))

    wb.close()

    long_df = pd.DataFrame(rows, columns=["raw_id", "date", "px_last", "px_volume"])
    long_df = long_df.join(sec_by_raw_id, on="raw_id")
    long_df = long_df.rename(columns={"FIGI": "figi", "EXPIRY": "expiry", "STRIKE": "strike"})
    long_df = long_df[["raw_id", "figi", "expiry", "strike", "date", "px_last", "px_volume"]]
    long_df["date"] = pd.to_datetime(long_df["date"])
    long_df["expiry"] = pd.to_datetime(long_df["expiry"])
    long_df["px_last"] = pd.to_numeric(long_df["px_last"], errors="coerce")
    long_df["px_volume"] = pd.to_numeric(long_df["px_volume"], errors="coerce")

    out_path = DATA_DIR / "processed" / f"{TICKER}_calls_close.parquet"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    long_df.to_parquet(out_path)

    print(f"Rows: {len(long_df):,}")
    print(f"Unique contracts: {long_df['raw_id'].nunique():,} (expect 3,797)")
    print(f"Date range: {long_df['date'].min().date()} -> {long_df['date'].max().date()}")
    print(f"Saved -> {out_path}")


if __name__ == "__main__":
    main()
