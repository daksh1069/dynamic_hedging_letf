"""
Stage 0a: extract every unique call-option BBG ID from
data/Excel5_OptionTickers_Final.xlsx (one sheet per
underlying, across the whole team's tickers) and write
excel_formula/Excel5b_UniqueTickers_ForDecode.xlsx, one sheet per
underlying with the deduplicated IDs in column A.

Output feeds excel_formula/scripts/02_decode_openfigi.py (once moved to
data/).

No API calls -- pure local file generation, free.
"""
import re
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"

INPUT_FILE  = DATA_DIR / "Excel5_OptionTickers_Final.xlsx"
OUTPUT_FILE = ROOT / "excel_formula" / "Excel5b_UniqueTickers_ForDecode.xlsx"

SKIP_SHEETS = {"INDEX", "HOW_TO_USE"}

# BBG ID pattern: BBG followed by alphanumerics, then a Bloomberg asset class suffix
BBG_PATTERN = re.compile(r"^BBG[A-Z0-9]{9,12}\s+(Equity|Curncy|Comdty|Index)$", re.IGNORECASE)

# ── COLORS ───────────────────────────────────────────────────────────────────
NAVY    = "1F3864"
BLUE    = "2E75B6"
GREEN   = "E2EFDA"
YELLOW  = "FFF2CC"
WHITE   = "FFFFFF"
LGREY   = "F2F2F2"

def fill(c): return PatternFill("solid", fgColor=c)
def hfont(size=9, bold=True, color="FFFFFF"):
    return Font(name="Arial", bold=bold, size=size, color=color)
def cfont(size=9, bold=False, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)
def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center")
def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# ── STEP 1: READ EXCEL 5 AND EXTRACT UNIQUE BBG IDs ─────────────────────────
print(f"Reading {INPUT_FILE} ...")
wb_in = load_workbook(INPUT_FILE, data_only=True)

sheet_ids = {}   # { sheet_name: [unique BBG IDs in order first seen] }
total_raw = 0

for sheet_name in wb_in.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue
    ws = wb_in[sheet_name]
    seen = set()
    ordered = []
    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value).strip() if cell.value is not None else ""
            if BBG_PATTERN.match(val) and val not in seen:
                seen.add(val)
                ordered.append(val)
                total_raw += 1
    sheet_ids[sheet_name] = ordered
    print(f"  {sheet_name:8s}: {len(ordered):>5,} unique BBG IDs")

print(f"\nTotal unique IDs across all sheets: {sum(len(v) for v in sheet_ids.values()):,}")

# ── STEP 2: BUILD OUTPUT EXCEL WITH BDP DECODE FORMULAS ──────────────────────
print(f"\nBuilding {OUTPUT_FILE} ...")
wb_out = Workbook()
wb_out.remove(wb_out.active)

# ── INDEX sheet ──────────────────────────────────────────────────────────────
idx = wb_out.create_sheet("INDEX")
idx.sheet_view.showGridLines = False

idx.merge_cells("A1:F1")
c = idx["A1"]
c.value = "Excel 5b — Unique Call Option IDs for BDP Decode"
c.font = hfont(size=13); c.fill = fill(NAVY); c.alignment = center()
idx.row_dimensions[1].height = 28

idx.merge_cells("A2:F2")
c = idx["A2"]
c.value = ("Open on Bloomberg Terminal → let BDP formulas refresh → "
           "select all → Paste Special Values → Save.  "
           "Then bring back to Python for moneyness filtering.")
c.font = Font(name="Arial", size=8, italic=True, color="FFFFFF")
c.fill = fill(BLUE); c.alignment = left()
idx.row_dimensions[2].height = 16

for col, h in enumerate(["#","Sheet (Underlying)","Unique BBG IDs","Next Step"], 1):
    c = idx.cell(row=4, column=col, value=h)
    c.font = hfont(9); c.fill = fill(BLUE); c.alignment = center()
idx.row_dimensions[4].height = 16

for i, (name, ids) in enumerate(sheet_ids.items()):
    row = i + 5
    bg = WHITE if i % 2 == 0 else LGREY
    for col, val in enumerate([i+1, name, len(ids), "Open on Bloomberg → refresh → paste values"], 1):
        c = idx.cell(row=row, column=col, value=val)
        c.font = cfont(9); c.fill = fill(bg)
        c.alignment = center() if col in [1,3] else left()
        c.border = thin()
    idx.row_dimensions[row].height = 14

for col, w in zip("ABCDF", [6, 20, 16, 40]):
    idx.column_dimensions[col].width = w

# ── HOW TO USE sheet ─────────────────────────────────────────────────────────
how = wb_out.create_sheet("HOW_TO_USE")
how.sheet_view.showGridLines = False
how.column_dimensions["A"].width = 12
how.column_dimensions["B"].width = 70

how.merge_cells("A1:B1")
c = how["A1"]
c.value = "HOW TO USE — Excel 5b"
c.font = hfont(13); c.fill = fill(NAVY); c.alignment = center()
how.row_dimensions[1].height = 28

steps = [
    ("PURPOSE",  "This file contains every unique call option BBG ID extracted from Excel 5, deduplicated per underlying. Use it to decode contract metadata (ticker, strike, expiry) via Bloomberg BDP."),
    ("STEP 1",   "Copy this file to the Bloomberg Terminal machine (USB / OneDrive / email)."),
    ("STEP 2",   "Open on Bloomberg Terminal. BDP formulas in columns B-E auto-populate with strike, expiry, put/call type, and human-readable ticker."),
    ("STEP 3",   "Once populated: select all → Copy → Paste Special → Values → Save as Excel5b_DECODED.xlsx"),
    ("STEP 4",   "Bring Excel5b_DECODED.xlsx back to your personal machine. Run the next Python script to filter by moneyness and maturity and build Excel 6."),
    ("COLUMNS",  "A: BBG Global ID  |  B: Human Ticker  |  C: Expiry Date  |  D: Strike Price  |  E: Put/Call (verify = C)"),
    ("⚠ NOTE",   "BDP may return #N/A for very old expired contracts with limited data. This is normal — those rows can be dropped in Python filtering."),
]
for i, (label, text) in enumerate(steps):
    row = i + 3
    c = how.cell(row=row, column=1, value=label)
    c.font = hfont(9); c.fill = fill(NAVY if "STEP" in label else BLUE if label in ["PURPOSE","COLUMNS"] else "C00000")
    c.alignment = center()
    c = how.cell(row=row, column=2, value=text)
    c.font = cfont(9); c.alignment = Alignment(vertical="center", wrap_text=True)
    how.row_dimensions[row].height = 40

# ── ONE SHEET PER UNDERLYING ─────────────────────────────────────────────────
for name, ids in sheet_ids.items():
    ws = wb_out.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"{name} — {len(ids):,} Unique Call Option IDs → BDP Decode"
    c.font = hfont(11); c.fill = fill(NAVY); c.alignment = left()
    ws.row_dimensions[1].height = 22

    # Sub-header
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = "Open on Bloomberg → formulas auto-fill → paste values → save → bring back to Python"
    c.font = Font(name="Arial", size=8, italic=True, color="FFFFFF")
    c.fill = fill(BLUE); c.alignment = left()
    ws.row_dimensions[2].height = 14

    # Column headers
    headers = ["BBG Global ID", "Ticker (BDP)", "Expiry Date (BDP)", "Strike Price (BDP)", "Put/Call (BDP)"]
    col_widths = [28, 28, 18, 16, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hfont(9); c.fill = fill(BLUE); c.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 16

    # Data rows
    for i, bbg_id in enumerate(ids):
        row = i + 4
        bg = WHITE if i % 2 == 0 else LGREY

        # Column A: BBG ID (value)
        c = ws.cell(row=row, column=1, value=bbg_id)
        c.font = Font(name="Courier New", size=8, color="000080")
        c.fill = fill(bg); c.alignment = left(); c.border = thin()

        # Columns B-E: BDP formulas
        fields = ["TICKER", "OPT_EXPIRE_DT", "OPT_STRIKE_PX", "OPT_PUT_CALL"]
        for col, field in enumerate(fields, 2):
            formula = f'=BDP(A{row},"{field}")'
            c = ws.cell(row=row, column=col, value=formula)
            c.font = cfont(9)
            c.fill = fill(GREEN if col == 3 else YELLOW if col == 4 else bg)
            c.alignment = center(); c.border = thin()

        ws.row_dimensions[row].height = 13

wb_out.save(OUTPUT_FILE)
print(f"\n✓ Saved: {OUTPUT_FILE}")
print("Next: copy to Bloomberg terminal → let BDP refresh → paste values → save → bring back for filtering.")
