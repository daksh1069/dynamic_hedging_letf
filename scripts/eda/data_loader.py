"""
Loaders for the raw Bloomberg / yfinance Excel exports in data/.

Every function returns a clean pandas DataFrame (or dict of DataFrames),
parsed directly from the raw .xlsx files -- no intermediate pipeline
artifacts required.
"""
import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"

NA_TOKENS = ["#N/A N/A", "#N/A Invalid Field", "#N/A Requesting Data..."]


def _clean_ohlcv(df: pd.DataFrame) -> pd.DataFrame:
    df = df.replace(NA_TOKENS, np.nan)
    df["Date"] = pd.to_datetime(df["Date"])
    for col in df.columns:
        if col != "Date":
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.dropna(subset=["Close"]).sort_values("Date").reset_index(drop=True)


def load_tsla_underlying() -> pd.DataFrame:
    """TSLA spot OHLCV + total return index + market cap, from Excel3_Underlying_Data.xlsx."""
    df = pd.read_excel(DATA_DIR / "Excel3_Underlying_Data.xlsx", sheet_name="TSLA", header=10)
    return _clean_ohlcv(df)


def load_tslt() -> pd.DataFrame:
    """TSLT (T-Rex 2X Long TSLA Daily Target ETF) OHLCV from Excel2_LETF_Data.xlsx.

    Data is NaN before TSLT's inception (2023-10-18); those rows are dropped.
    """
    df = pd.read_excel(DATA_DIR / "Excel2_LETF_Data.xlsx", sheet_name="TSLT", header=17)
    return _clean_ohlcv(df)


def load_tsll() -> pd.DataFrame:
    """TSLL (Direxion Daily TSLA Bull 2X Shares) OHLCV from data/TSLL_ohlcv.xlsx (yfinance)."""
    df = pd.read_excel(DATA_DIR / "TSLL_ohlcv.xlsx", sheet_name="TSLL")
    df["Date"] = pd.to_datetime(df["Date"])
    return df.sort_values("Date").reset_index(drop=True)


def load_benchmarks() -> dict:
    """SPY, QQQ, VIX, and 3M T-bill yield from Benchmarks.xlsx (multi-block layout)."""
    raw = pd.read_excel(DATA_DIR / "Benchmarks.xlsx", sheet_name="Sheet1", header=8)

    def block(col_idx, names):
        sub = raw.iloc[:, col_idx].copy()
        sub.columns = names
        sub["Date"] = pd.to_datetime(sub["Date"], unit="D", origin="1899-12-30")
        for c in names[1:]:
            sub[c] = pd.to_numeric(sub[c], errors="coerce")
        return sub.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)

    return {
        "SPY":    block(range(0, 7),   ["Date", "Open", "High", "Low", "Close", "TRI", "Volume"]),
        "QQQ":    block(range(8, 15),  ["Date", "Open", "High", "Low", "Close", "TRI", "Volume"]),
        "VIX":    block(range(16, 21), ["Date", "Open", "High", "Low", "Close"]),
        "USGG3M": block(range(22, 24), ["Date", "Yield_3M"]),
    }


def load_tsla_calls(use_cache: bool = True) -> pd.DataFrame:
    """Long-format TSLA call option closing prices.

    Parses data/TSLA_calls_PXLAST_full_filled.xlsx (Securities sheet for
    contract metadata + Batch_NN sheets for [Date | PX_LAST | PX_VOLUME]
    blocks, one block per contract). Caches the parsed result to
    data/processed/TSLA_calls_close.parquet for fast reuse.

    Columns: raw_id, figi, expiry, strike, date, px_last, px_volume
    """
    cache_path = DATA_DIR / "processed" / "TSLA_calls_close.parquet"
    if use_cache and cache_path.exists():
        return pd.read_parquet(cache_path)

    path = DATA_DIR / "TSLA_calls_PXLAST_full_filled.xlsx"
    sec = pd.read_excel(path, sheet_name="Securities")
    sec_by_raw_id = sec.set_index("RAW_ID")[["FIGI", "EXPIRY", "STRIKE"]]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    BLOCK_WIDTH = 4
    rows = []
    for name in wb.sheetnames:
        if not name.startswith("Batch"):
            continue
        ws = wb[name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        max_col = max(len(r) for r in sheet_rows)
        n_blocks = (max_col + BLOCK_WIDTH - 1) // BLOCK_WIDTH

        for b in range(n_blocks):
            c0 = b * BLOCK_WIDTH
            label = sheet_rows[0][c0] if c0 < len(sheet_rows[0]) else None
            if label is None:
                continue
            raw_id = label.split(" (")[0].strip()

            for r in sheet_rows[2:]:
                if c0 + 1 >= len(r):
                    continue
                date_val = r[c0]
                px_last = r[c0 + 1]
                px_vol = r[c0 + 2] if c0 + 2 < len(r) else None
                if not isinstance(date_val, datetime.datetime):
                    continue
                rows.append((raw_id, date_val, px_last, px_vol))
    wb.close()

    long_df = pd.DataFrame(rows, columns=["raw_id", "date", "px_last", "px_volume"])
    long_df = long_df.join(sec_by_raw_id, on="raw_id")
    long_df = long_df.rename(columns={"FIGI": "figi", "EXPIRY": "expiry", "STRIKE": "strike"})
    long_df = long_df[["raw_id", "figi", "expiry", "strike", "date", "px_last", "px_volume"]]
    long_df["date"] = pd.to_datetime(long_df["date"])
    long_df["expiry"] = pd.to_datetime(long_df["expiry"])
    long_df["px_last"] = pd.to_numeric(long_df["px_last"], errors="coerce")
    long_df["px_volume"] = pd.to_numeric(long_df["px_volume"], errors="coerce")

    if use_cache:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        long_df.to_parquet(cache_path)

    return long_df


def load_tsll_calls() -> pd.DataFrame:
    """Long-format TSLL call option closing prices.

    Reads the pre-parsed cache at data/processed/TSLL_calls_close.parquet.
    Columns: raw_id, figi, expiry, strike, date, px_last, px_volume
    """
    return pd.read_parquet(DATA_DIR / "processed" / "TSLL_calls_close.parquet")


if __name__ == "__main__":
    tsla = load_tsla_underlying()
    tslt = load_tslt()
    tsll = load_tsll()
    bench = load_benchmarks()
    calls = load_tsla_calls()

    print(f"TSLA : {len(tsla):,} rows  {tsla['Date'].min().date()} -> {tsla['Date'].max().date()}")
    print(f"TSLT : {len(tslt):,} rows  {tslt['Date'].min().date()} -> {tslt['Date'].max().date()}")
    print(f"TSLL : {len(tsll):,} rows  {tsll['Date'].min().date()} -> {tsll['Date'].max().date()}")
    for name, df in bench.items():
        print(f"{name:6s}: {len(df):,} rows  {df['Date'].min().date()} -> {df['Date'].max().date()}")
    print(f"CALLS: {len(calls):,} rows  {calls['raw_id'].nunique():,} contracts  "
          f"{calls['date'].min().date()} -> {calls['date'].max().date()}")
    tsll_calls = load_tsll_calls()
    print(f"TSLL_CALLS: {len(tsll_calls):,} rows  {tsll_calls['raw_id'].nunique():,} contracts  "
          f"{tsll_calls['date'].min().date()} -> {tsll_calls['date'].max().date()}")
